import pandas as pd
import sys
import json
from openpyxl import load_workbook

def append_to_excel(model_name, data):
    file_name = f"{model_name}_Attendance.xlsx"
    
    try:
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = pd.ExcelWriter(file_name, engine='openpyxl')
            df = pd.DataFrame(columns=['Sr No', 'PRN', 'Attendance'])
            df.to_excel(workbook, index=False)
            workbook.save()
            workbook = load_workbook(file_name)
            sheet = workbook.active

        users = pd.DataFrame(data['users'])
        prns = pd.DataFrame(data['prns'])
        
        for prn in prns['prn']:
            if prn not in sheet['B']:
                new_row = [sheet.max_row + 1, prn, '']
                sheet.append(new_row)
        
        today = pd.Timestamp.today().strftime('%Y-%m-%d')
        if today not in sheet.columns:
            sheet.insert_cols(sheet.max_column + 1)
            sheet.cell(row=1, column=sheet.max_column).value = today
            
        for user in users['userId']:
            if user in sheet['B']:
                row_idx = sheet['B'].index(user) + 2
                sheet.cell(row=row_idx, column=sheet.max_column).value = user
                
        workbook.save(file_name)
        print(f"Data successfully appended to {file_name}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    model_name = sys.argv[1]
    data = json.loads(sys.argv[2])
    append_to_excel(model_name, data)